from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# creat a work object
wb = Workbook()

# create an active worksheet
ws = wb.active

# load existng spredsheet

wb = load_workbook('text.xlsx')